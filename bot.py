import os
import re
import logging
import tempfile
import subprocess
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from telegram import Update
from telegram.ext import ApplicationBuilder, MessageHandler, CommandHandler, filters, ContextTypes

TOKEN = os.environ.get("BOT_TOKEN", "YOUR_TOKEN_HERE")
logging.basicConfig(level=logging.INFO)

def decode_text(s):
    if isinstance(s, bytes):
        for enc in ["windows-1251", "utf-8", "cp1252"]:
            try:
                return s.decode(enc)
            except:
                continue
    return str(s) if s else ""

def parse_html_xls(filepath):
    with open(filepath, "rb") as f:
        raw = f.read()
    for enc in ["windows-1251", "utf-8", "cp1252"]:
        try:
            content = raw.decode(enc)
            break
        except:
            content = raw.decode("windows-1251", errors="replace")

    from html.parser import HTMLParser
    class TableParser(HTMLParser):
        def __init__(self):
            super().__init__()
            self.tables = []
            self.current_table = []
            self.current_row = []
            self.current_cell = ""
            self.in_cell = False
            self.in_table = False
        def handle_starttag(self, tag, attrs):
            if tag == "table":
                self.in_table = True
                self.current_table = []
            elif tag in ("tr",):
                self.current_row = []
            elif tag in ("td", "th"):
                self.in_cell = True
                self.current_cell = ""
        def handle_endtag(self, tag):
            if tag == "table":
                self.tables.append(self.current_table)
                self.in_table = False
            elif tag == "tr":
                if self.current_row:
                    self.current_table.append(self.current_row)
            elif tag in ("td", "th"):
                self.current_row.append(self.current_cell.strip())
                self.in_cell = False
        def handle_data(self, data):
            if self.in_cell:
                self.current_cell += data

    parser = TableParser()
    parser.feed(content)
    return parser.tables, content

def extract_num_from_str(s):
    s = str(s).replace(" ", "").replace("\xa0", "").replace(",", ".")
    m = re.search(r"[\d]+\.?\d*", s)
    if m:
        try:
            return float(m.group())
        except:
            return 0
    return 0

def clean_counterparty(s):
    s = str(s).strip()
    parts = s.split("/")
    if len(parts) >= 3:
        name = parts[2].strip()
    else:
        name = s
    name = re.sub(r"^\d+\s*", "", name).strip()
    if len(name) > 50:
        name = name[:50] + "..."
    return name if name else "Noma'lum"

def parse_file(filepath, fname):
    ext = fname.lower().split(".")[-1]
    if ext in ("xlsb", "xlsx", "xls"):
        with open(filepath, "rb") as f:
            header = f.read(8)
        is_html = header[:5] in (b"<html", b"<!DOC", b"<HTML") or b"<html" in header.lower()[:100]
        if is_html or ext == "xls":
            with open(filepath, "rb") as f:
                raw = f.read()
            for enc in ["windows-1251", "utf-8"]:
                try:
                    content = raw.decode(enc)
                    break
                except:
                    content = raw.decode("windows-1251", errors="replace")
            if "<html" in content.lower():
                return parse_html_content(content)
        out_dir = tempfile.mkdtemp()
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "xlsx", "--outdir", out_dir, filepath],
            capture_output=True, text=True, timeout=30
        )
        xlsx_files = [f for f in os.listdir(out_dir) if f.endswith(".xlsx")]
        if not xlsx_files:
            raise Exception("Fayl o'qilmadi")
        df = pd.read_excel(os.path.join(out_dir, xlsx_files[0]), sheet_name=0, header=None)
        return parse_dataframe(df)

def parse_html_content(content):
    from html.parser import HTMLParser
    class TP(HTMLParser):
        def __init__(self):
            super().__init__()
            self.tables = []
            self.cur_table = []
            self.cur_row = []
            self.cur_cell = ""
            self.in_cell = False
        def handle_starttag(self, tag, attrs):
            if tag == "table":
                self.cur_table = []
            elif tag == "tr":
                self.cur_row = []
            elif tag in ("td", "th"):
                self.in_cell = True
                self.cur_cell = ""
            elif tag == "br" and self.in_cell:
                self.cur_cell += " "
        def handle_endtag(self, tag):
            if tag == "table":
                self.tables.append(self.cur_table)
            elif tag == "tr":
                if self.cur_row:
                    self.cur_table.append(self.cur_row)
            elif tag in ("td", "th"):
                self.cur_row.append(self.cur_cell.strip())
                self.in_cell = False
        def handle_data(self, data):
            if self.in_cell:
                self.cur_cell += data
        def handle_entityref(self, name):
            if self.in_cell and name == "nbsp":
                self.cur_cell += " "

    p = TP()
    p.feed(content)

    # 1-jadval: bank info
    # 2-jadval: tranzaksiyalar
    info = {}
    txs = []
    total_debit = 0
    total_credit = 0

    for table in p.tables:
        for row in table:
            if not row:
                continue
            row0 = str(row[0])
            # Bank nomi
            if "БАНК" in row0.upper() or "BANK" in row0.upper() or "00873" in row0 or "00842" in row0:
                if "bank" not in info:
                    info["bank"] = row0.strip()
            # Davr
            if "Сведения" in row0 or "c 0" in row0:
                info["period"] = row0.strip()
            # Schet va korxona
            if "Счет" in row0 or "Cчет" in row0 or "счёт" in row0.lower():
                info["account"] = row0.strip()
                if len(row) > 1:
                    info["company"] = str(row[1]).strip()
                if len(row) > 2:
                    info["inn"] = str(row[2]).strip()
            # Qoldiqlar
            if "Остаток на начало" in row0 or "начало" in row0:
                info["bal_start_str"] = row0
                info["bal_start"] = extract_num_from_str(row0.split(":")[-1])
                if len(row) > 1:
                    info["bal_end_str"] = str(row[1])
                    info["bal_end"] = extract_num_from_str(str(row[1]).split(":")[-1])
            # Jami (Итоговый)
            if "Итогов" in row0 or "Итого" in row0:
                if len(row) >= 7:
                    try:
                        td = extract_num_from_str(row[5])
                        tc = extract_num_from_str(row[6])
                        if td > 0 or tc > 0:
                            total_debit = td
                            total_credit = tc
                    except:
                        pass
                continue
            # Tranzaksiya qatori (sana formatida)
            if len(row) >= 7:
                date_str = str(row[0]).strip()
                m = re.match(r"(\d{2}\.\d{2}\.\d{4})", date_str)
                if m:
                    date = m.group(1)
                    time_m = re.search(r"(\d{2}:\d{2}:\d{2})", date_str)
                    time = time_m.group(1)[:5] if time_m else ""
                    counterparty_raw = str(row[1]).strip()
                    counterparty = clean_counterparty(counterparty_raw)
                    doc = str(row[2]).strip()
                    debit_str = str(row[5]).strip().replace("\xa0", "").replace(" ", "")
                    credit_str = str(row[6]).strip().replace("\xa0", "").replace(" ", "")
                    debit = extract_num_from_str(debit_str) if debit_str and debit_str != "&nbsp;" else None
                    credit = extract_num_from_str(credit_str) if credit_str and credit_str != "&nbsp;" else None
                    desc = str(row[7]).strip() if len(row) > 7 else ""
                    if desc.startswith("00634") or desc.startswith("00668") or desc.startswith("00508") or desc.startswith("00599") or desc.startswith("1900") or desc.startswith("0200"):
                        desc = re.sub(r"^0\d{4}\s*", "", desc).strip()
                    txs.append({
                        "date": date, "time": time, "doc": doc,
                        "counterparty": counterparty,
                        "debit": debit if debit and debit > 0 else None,
                        "credit": credit if credit and credit > 0 else None,
                        "desc": desc[:100]
                    })

    if total_debit == 0:
        total_debit = sum(tx["debit"] for tx in txs if tx["debit"])
    if total_credit == 0:
        total_credit = sum(tx["credit"] for tx in txs if tx["credit"])

    return info, txs, total_debit, total_credit

def parse_dataframe(df):
    info = {}
    try:
        info["bank"] = str(df.iloc[0, 0]).strip()
        info["period"] = str(df.iloc[1, 0]).strip()
        info["account"] = str(df.iloc[2, 0]).replace("\xa0", " ").strip()
        info["bal_start"] = extract_num_from_str(str(df.iloc[3, 0]))
        info["bal_end"] = extract_num_from_str(str(df.iloc[3, 1]))
        info["company"] = str(df.iloc[3, 7]).strip() if pd.notna(df.iloc[3, 7]) else ""
    except:
        pass

    txs = []
    total_debit = 0
    total_credit = 0
    header_row = None
    for i, r in df.iterrows():
        if str(r.iloc[0]).strip() in ("\u0414\u0430\u0442\u0430", "Dата", "Data"):
            header_row = i
            break
    if header_row:
        for i in range(header_row + 1, len(df)):
            r = df.iloc[i]
            date_val = r.iloc[0]
            if pd.isna(date_val):
                continue
            if "\u0418\u0442\u043e\u0433" in str(date_val):
                try:
                    total_debit = float(r.iloc[5]) if pd.notna(r.iloc[5]) else 0
                    total_credit = float(r.iloc[6]) if pd.notna(r.iloc[6]) else 0
                except:
                    pass
                continue
            try:
                date_fmt = date_val.strftime("%d.%m.%Y") if hasattr(date_val, "strftime") else str(date_val)[:10]
                time_fmt = date_val.strftime("%H:%M") if hasattr(date_val, "strftime") else ""
                doc = str(int(r.iloc[2])) if pd.notna(r.iloc[2]) else ""
                counterparty = clean_counterparty(str(r.iloc[1]) if pd.notna(r.iloc[1]) else "")
                debit = float(r.iloc[5]) if pd.notna(r.iloc[5]) else None
                credit = float(r.iloc[6]) if pd.notna(r.iloc[6]) else None
                desc = str(r.iloc[7]).strip() if pd.notna(r.iloc[7]) else ""
                txs.append({"date": date_fmt, "time": time_fmt, "doc": doc,
                            "counterparty": counterparty, "debit": debit, "credit": credit, "desc": desc[:100]})
            except:
                continue
    if total_debit == 0:
        total_debit = sum(tx["debit"] for tx in txs if tx["debit"])
    if total_credit == 0:
        total_credit = sum(tx["credit"] for tx in txs if tx["credit"])
    return info, txs, total_debit, total_credit

def build_summary(txs):
    debit_by = defaultdict(lambda: {"sum": 0, "count": 0})
    credit_by = defaultdict(lambda: {"sum": 0, "count": 0})
    for tx in txs:
        name = tx["counterparty"]
        if tx["debit"]:
            debit_by[name]["sum"] += tx["debit"]
            debit_by[name]["count"] += 1
        if tx["credit"]:
            credit_by[name]["sum"] += tx["credit"]
            credit_by[name]["count"] += 1
    return debit_by, credit_by

def thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)

def sty(cell, bold=False, bg=None, color="000000", size=10, align="left"):
    cell.font = Font(bold=bold, color=color, size=size, name="Arial")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = thin_border()

def create_excel(info, txs, total_debit, total_credit, output_path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Kochirma"
    DARK = "1B3A6B"
    LBLUE = "D6E4F7"
    RED = "C0392B"
    GREEN = "1E8449"
    WHITE = "FFFFFF"
    YELLOW = "FFF3CD"

    bank = info.get("bank", "Bank")
    period = info.get("period", "")
    account = info.get("account", "")
    company = info.get("company", "")
    bal_start = info.get("bal_start", 0)
    bal_end = info.get("bal_end", 0)

    row = 1
    ws1.merge_cells("A%d:H%d" % (row, row))
    c = ws1.cell(row=row, column=1, value=bank)
    sty(c, bold=True, bg=DARK, color=WHITE, size=12, align="center")
    ws1.row_dimensions[row].height = 26
    row += 1

    ws1.merge_cells("A%d:H%d" % (row, row))
    c = ws1.cell(row=row, column=1, value=period)
    sty(c, bg=LBLUE, size=10, align="center")
    ws1.row_dimensions[row].height = 20
    row += 1

    if account:
        ws1.merge_cells("A%d:H%d" % (row, row))
        c = ws1.cell(row=row, column=1, value=account)
        sty(c, size=9, align="center")
        ws1.row_dimensions[row].height = 18
        row += 1

    if company and company != "nan":
        ws1.merge_cells("A%d:H%d" % (row, row))
        c = ws1.cell(row=row, column=1, value=company)
        sty(c, bold=True, bg=YELLOW, size=10, align="center")
        ws1.row_dimensions[row].height = 20
        row += 1

    row += 1
    for i, (lbl, val, col) in enumerate([
        ("Boshlanish qoldighi", bal_start, DARK),
        ("Tugash qoldighi", bal_end, RED if bal_end < bal_start else GREEN)
    ]):
        c1 = get_column_letter(1 + i * 4)
        c2 = get_column_letter(3 + i * 4)
        ws1.merge_cells("%s%d:%s%d" % (c1, row, c2, row))
        c = ws1.cell(row=row, column=1 + i * 4, value=lbl)
        sty(c, bold=True, bg=col, color=WHITE, align="center")
        ws1.row_dimensions[row].height = 20
        ws1.merge_cells("%s%d:%s%d" % (c1, row + 1, c2, row + 1))
        c2cell = ws1.cell(row=row + 1, column=1 + i * 4, value=val)
        sty(c2cell, bold=True, bg=LBLUE, align="center", size=13)
        c2cell.number_format = "#,##0.00"
        ws1.row_dimensions[row + 1].height = 24
    row += 3

    for col_i, h in enumerate(["Sana", "Vaqt", "Hujjat", "Kontragent", "Chiqim", "Kirim", "Izoh"], 1):
        c = ws1.cell(row=row, column=col_i, value=h)
        sty(c, bold=True, bg=DARK, color=WHITE, align="center")
    ws1.row_dimensions[row].height = 20
    row += 1

    for tx in txs:
        bg = "FFE8E8" if tx["debit"] else "E8FFE8"
        vals = [tx["date"], tx["time"], tx["doc"], tx["counterparty"], tx["debit"], tx["credit"], tx["desc"]]
        for col_i, val in enumerate(vals, 1):
            c = ws1.cell(row=row, column=col_i, value=val)
            sty(c, bg=bg, size=9)
            if col_i == 5 and val:
                sty(c, bold=True, bg=bg, color=RED, align="right", size=10)
                c.number_format = "#,##0.00"
            if col_i == 6 and val:
                sty(c, bold=True, bg=bg, color=GREEN, align="right", size=10)
                c.number_format = "#,##0.00"
        ws1.row_dimensions[row].height = 16
        row += 1

    row += 1
    for lbl, val, col, col_i in [
        ("Jami chiqim:", total_debit, RED, 5),
        ("Jami kirim:", total_credit, GREEN, 6)
    ]:
        ws1.merge_cells("A%d:D%d" % (row, row))
        c = ws1.cell(row=row, column=1, value=lbl)
        sty(c, bold=True, align="right", size=11)
        c2 = ws1.cell(row=row, column=col_i, value=val)
        sty(c2, bold=True, color=col, align="right", size=12)
        c2.number_format = "#,##0.00"
        ws1.row_dimensions[row].height = 22
        row += 1

    for i, w in enumerate([12, 7, 13, 28, 16, 16, 40], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Firmalar")
    debit_by, credit_by = build_summary(txs)
    all_names = set(list(debit_by.keys()) + list(credit_by.keys()))

    row2 = 1
    ws2.merge_cells("A%d:E%d" % (row2, row2))
    c = ws2.cell(row=row2, column=1, value="FIRMALAR BO'YICHA JAMLAMA")
    sty(c, bold=True, bg=DARK, color=WHITE, size=13, align="center")
    ws2.row_dimensions[row2].height = 28
    row2 += 1

    ws2.merge_cells("A%d:E%d" % (row2, row2))
    c = ws2.cell(row=row2, column=1, value=period)
    sty(c, bg=LBLUE, size=10, align="center")
    ws2.row_dimensions[row2].height = 20
    row2 += 2

    for col_i, h in enumerate(["Firma nomi", "Chiqim soni", "Jami chiqim", "Kirim soni", "Jami kirim"], 1):
        c = ws2.cell(row=row2, column=col_i, value=h)
        sty(c, bold=True, bg=DARK, color=WHITE, align="center")
    ws2.row_dimensions[row2].height = 22
    row2 += 1

    for name in sorted(all_names):
        d = debit_by.get(name, {"sum": 0, "count": 0})
        cr = credit_by.get(name, {"sum": 0, "count": 0})
        bg = "FFF0F0" if d["sum"] > cr["sum"] else "F0FFF0" if cr["sum"] > d["sum"] else "FFFFFF"
        for col_i, val in enumerate([name, d["count"] or "", d["sum"] or "", cr["count"] or "", cr["sum"] or ""], 1):
            c = ws2.cell(row=row2, column=col_i, value=val)
            sty(c, bg=bg, size=10)
            if col_i == 3 and val:
                sty(c, bold=True, bg=bg, color=RED, align="right", size=11)
                c.number_format = "#,##0.00"
            if col_i == 5 and val:
                sty(c, bold=True, bg=bg, color=GREEN, align="right", size=11)
                c.number_format = "#,##0.00"
        ws2.row_dimensions[row2].height = 20
        row2 += 1

    for col, w in zip(["A", "B", "C", "D", "E"], [35, 13, 18, 13, 18]):
        ws2.column_dimensions[col].width = w

    wb.save(output_path)

def build_message(info, txs, total_debit, total_credit):
    debit_by, credit_by = build_summary(txs)
    all_names = set(list(debit_by.keys()) + list(credit_by.keys()))
    bank = info.get("bank", "Bank")
    period = info.get("period", "")
    company = info.get("company", "")
    bal_start = info.get("bal_start", 0)
    bal_end = info.get("bal_end", 0)

    msg = []
    msg.append("*" + bank[:60] + "*")
    if period:
        msg.append(period[:80])
    if company and company != "nan":
        msg.append("Korxona: *" + company[:50] + "*")
    msg.append("")
    msg.append("Boshlanish: `%s` sum" % "{:,.0f}".format(bal_start))
    msg.append("Tugash: `%s` sum" % "{:,.0f}".format(bal_end))
    msg.append("")
    msg.append("Jami chiqim: *%s sum*" % "{:,.0f}".format(total_debit))
    msg.append("Jami kirim: *%s sum*" % "{:,.0f}".format(total_credit))
    msg.append("")
    msg.append("*FIRMALAR (%d ta):*" % len(all_names))
    for name in sorted(all_names):
        d = debit_by.get(name, {"sum": 0, "count": 0})
        cr = credit_by.get(name, {"sum": 0, "count": 0})
        msg.append("")
        msg.append("*" + name + "*")
        if d["sum"]:
            msg.append("  Chiqim: `%s` sum (%d ta)" % ("{:,.0f}".format(d["sum"]), d["count"]))
        if cr["sum"]:
            msg.append("  Kirim: `%s` sum (%d ta)" % ("{:,.0f}".format(cr["sum"]), cr["count"]))
    return "\n".join(msg)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Salom!\n\n"
        ".xlsb, .xlsx yoki .xls bank kochirmasini yuboring.\n"
        "Men xabar va Excel hisobot qaytaraman."
    )

async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    if not doc:
        return
    fname = doc.file_name or ""
    if not (fname.lower().endswith(".xlsb") or fname.lower().endswith(".xlsx") or fname.lower().endswith(".xls")):
        await update.message.reply_text("Iltimos, .xlsb, .xlsx yoki .xls fayl yuboring.")
        return
    await update.message.reply_text("Fayl qayta ishlanmoqda...")
    with tempfile.TemporaryDirectory() as tmpdir:
        fpath = os.path.join(tmpdir, fname)
        file = await doc.get_file()
        await file.download_to_drive(fpath)
        try:
            result = parse_file(fpath, fname)
            info, txs, total_debit, total_credit = result
            msg = build_message(info, txs, total_debit, total_credit)
            if len(msg) > 4000:
                msg = msg[:4000] + "\n..."
            await update.message.reply_text(msg, parse_mode="Markdown")
            out_path = os.path.join(tmpdir, "hisobot.xlsx")
            create_excel(info, txs, total_debit, total_credit, out_path)
            with open(out_path, "rb") as f:
                await update.message.reply_document(
                    document=f,
                    filename="bank_hisoboti.xlsx",
                    caption="Excel hisoboti tayyor! (%d ta tranzaksiya)" % len(txs)
                )
        except Exception as e:
            await update.message.reply_text("Xato: " + str(e))

async def handle_other(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Iltimos, .xlsb, .xlsx yoki .xls fayl yuboring.")

def main():
    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_file))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_other))
    print("Bot ishlamoqda...")
    app.run_polling()

if __name__ == "__main__":
    main()
